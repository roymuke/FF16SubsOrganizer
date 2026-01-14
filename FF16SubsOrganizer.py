import os, argparse
import xml.etree.ElementTree as ET

def get_ids():
    import json
    with open("IDs.json","r",encoding="utf-8") as f:
        jsonIds = json.load(f)
    return jsonIds["characters"], jsonIds["subtitleID"]

def read_texts(xml_path):
    try:
        with open(xml_path, "r", encoding="utf-8") as f:
            content = f.read()
        root = ET.fromstring(content)
        text_contents = root.find("TextContents")
        if text_contents is None:
            return []
        result = []
        for text_content in text_contents.findall("TextContent"):
            content_id = text_content.get("ID", "")
            message = text_content.findtext("Message", default="").strip()
            chara_id = text_content.get("Unknown2", "")
            subtype = text_content.get("Unknown3", "")
            result.append((content_id, chara_id, subtype, message))
        return result
    except Exception as e:
        print(f" \033[91m[ERROR]\033[00m Error reading {xml_path}: {e}")
        return []

def fix_xml_fields(root):
    text_contents = root.find("TextContents")
    if text_contents is None:
        return
    for text_content in text_contents.findall("TextContent"):
        message_elem = text_content.find("Message")
        if message_elem is None:
            message_elem = ET.SubElement(text_content, "Message")
        if message_elem.text is None:
            message_elem.text = ""
        voice_elem = text_content.find("Voice")
        if voice_elem is None:
            voice_elem = ET.SubElement(text_content, "Voice")
        if voice_elem.text is None:
            voice_elem.text = ""
        string_elem = text_content.find("String")
        if string_elem is None:
            string_elem = ET.SubElement(text_content, "String")
        if string_elem.text is None:
            string_elem.text = ""

def write_xml(tree, path):
    root = tree.getroot()
    fix_xml_fields(root)
    xml_body = ET.tostring(root, encoding="unicode", method="xml")
    xml_content = '<?xml version="1.0" encoding="utf-16"?>\r\n' + xml_body
    with open(path, "w", encoding="utf-8", newline="") as f:
        f.write(xml_content)

def collect_table(lang_root, jap_root):
    table_rows = []
    characters, subtitleID = get_ids()
    for root_dir, _, files in os.walk(lang_root):
        for file in files:
            if not file.endswith(".xml"):
                continue
            lang_path = os.path.join(root_dir, file)
            rel_path = os.path.relpath(lang_path, lang_root)
            subdir = os.path.basename(os.path.dirname(rel_path))
            jap_path = os.path.join(jap_root, rel_path)
            if not os.path.exists(jap_path):
                print(f" \033[38;5;214m[WARNING]\033[00m Japanese path not found: {jap_path}")
                continue
            lang_data = read_texts(lang_path)
            jap_data = read_texts(jap_path)
            filename = os.path.splitext(file)[0]
            if filename.endswith(".pzd"):
                filename = filename[:-4]
            for idx, (id_msg, chara_id, subtype, en_msg) in enumerate(lang_data):
                jp_msg = jap_data[idx][3] if idx < len(jap_data) else ""
                chara_name = characters.get(chara_id, "")
                sub_type = subtitleID.get(subtype, "")
                table_rows.append((subdir, filename, id_msg, sub_type, chara_name, chara_id, en_msg, jp_msg))
    return table_rows

# Command: to-xlsx
def export_xlsx(table_rows, output, verbose):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from collections import Counter, defaultdict
    rows_by_subdir = defaultdict(list); stats = list(); batch_filename = list()
    print("> Generating file...")
    try:
        for row in table_rows:
            subdir = row[0]; rows_by_subdir[subdir].append(row)
        wb = Workbook()
        stats_ws = wb.active; stats_ws.title = "STATS"
        len_sheets = 10 + len(wb.sheetnames)
        for subdir, rows in rows_by_subdir.items():
            if verbose:
                print(f" \033[38;5;75m[INFO]\033[00m Processing: {subdir}")
            stats.append([subdir,"!D2:D"+ str(len(rows)+1),"!I2:I"+ str(len(rows)+1)])
            sheet_name = subdir[:31] if len(subdir) <= 31 else subdir[:28] + "..."
            sheet_name = sheet_name.replace("/", "_").replace("\\", "_").replace("[", "_").replace("]", "_").replace("*", "_").replace("?", "_").replace(":", "_")
            ws = wb.create_sheet(title=sheet_name)
            ws.append(["Folder", "Filename", "ID", "Sub Type", "Character", "Character ID", "Original Text", "Japanese", "Retranslation"])
            ws.freeze_panes = "A2"
            ws.column_dimensions["G"].width = 45; ws.column_dimensions["H"].width = 60; ws.column_dimensions["I"].width = 59; ws.column_dimensions["J"].width = 30
            ws.sheet_view.zoomScale = 80
            for subdir_name, filename, msg_id, subtype, chara_name, chara_id, en_text, jp_text in rows:
                ws.append([subdir_name, filename, msg_id, subtype, chara_name, chara_id, en_text, jp_text, ""])
                batch_filename.append(filename)
            ws.column_dimensions["B"].width = (len(str(ws["B2"].value)) + 0.5) * 1.1207692307692307
            row_filename_len = Counter(batch_filename); start_row = 2
            for i, count in enumerate(row_filename_len.items()):
                if i % 2 == 0:
                    for j in range(start_row, start_row + count[1]):
                        for col_num in range(1,10):
                            cell = ws.cell(row=j, column=col_num)
                            cell.fill = PatternFill(fill_type="solid", start_color="FFF2F2F2")
                else:
                    start_row = start_row + count[1]
                    continue
                start_row = start_row + count[1]
            batch_filename = list()
        stats_ws["A1"] = "FFXVI Subtitle Translation Progress Sheet"; stats_ws["A1"].font = Font(size="22"); stats_ws.merge_cells("A1:D1")
        stats_ws["A2"] = "Made with FF16SubsOrganizer"; stats_ws["A2"].font = Font(size="10"); stats_ws["A2"].alignment = Alignment(horizontal="right"); stats_ws.merge_cells("A2:D2")
        stats_ws.column_dimensions["A"].width = 12; stats_ws.column_dimensions["C"].width = 10; stats_ws.column_dimensions["D"].width = 50; len_sheets = 9 + len(wb.sheetnames)
        progress = [['Subtitle' 'Type','Lines','Translated','Progress'],['Normal',f'SUMPRODUCT(COUNTIF(INDIRECT(A11:A{len_sheets}&B11:B{len_sheets});"Normal"))',f'B5-SUMPRODUCT(COUNTIFS(INDIRECT(A11:A{len_sheets}&B11:B{len_sheets});"Normal";INDIRECT(A11:A{len_sheets}&C11:C{len_sheets});""))','C5/B5'],['SFX',f'SUMPRODUCT(COUNTIF(INDIRECT(A11:A{len_sheets}&B11:B{len_sheets});"SFX"))',f'B6-SUMPRODUCT(COUNTIFS(INDIRECT(A11:A{len_sheets}&B11:B{len_sheets});"SFX";INDIRECT(A11:A{len_sheets}&C11:C{len_sheets});""))','C6/B6'],['Hidden',f'SUMPRODUCT(COUNTIF(INDIRECT(A11:A{len_sheets}&B11:B{len_sheets});"Hidden"))',f'B7-SUMPRODUCT(COUNTIFS(INDIRECT(A11:A{len_sheets}&B11:B{len_sheets});"Hidden";INDIRECT(A11:A{len_sheets}&C11:C{len_sheets});""))','C7/B7'],['TOTAL','SUM(B5:B7)','SUM(C5:C7)','C8/B8']]
        stats_ws["A8"].font = stats_ws["B8"].font = stats_ws["C8"].font = stats_ws["D8"].font = Font(bold=True)
        i, j = 4, 11
        for _, row_data in enumerate(progress):
            for columna, valor in enumerate(row_data):
                stats_ws.cell(row=i, column=columna + 1, value=valor)
            i += 1
        for _, row_data in enumerate(stats):
            for columna, valor in enumerate(row_data):
                stats_ws.cell(row=j, column=columna + 1, value=valor)
            j += 1
        for row_num in range(11, len_sheets + 1):
            stats_ws.row_dimensions[row_num].hidden = True
        wb.save(output)
        print(f" \033[38;5;76m[DONE]\033[00m XLSX file generated in: \033[48;5;235m{output}\033[00m")
        print(f" \033[38;5;81m[INSTRUCTION] Edit the 'Retranslation' column (I) on each sheet. Once done, use 'edit-xml' to apply changes.\033[00m")
    except Exception as e:
        print(f" \033[91m[ERROR]\033[00m Couldn't generate file: {e}")

# Command: edit-xml
def edit_xml(xlsx_path, col_reference, lang_root, verbose):
    from openpyxl import load_workbook
    from openpyxl.utils import column_index_from_string
    from html import unescape
    try:
        wb = load_workbook(xlsx_path)
        all_sheets = {}
        for sheet_name in wb.sheetnames[1:]:
            sheet = wb[sheet_name]
            sheet_data = []
            for row in sheet.iter_rows(min_row=2,values_only=True):
                sheet_data.append(list(row))
            all_sheets[sheet_name] = sheet_data
        col_letter = "".join(filter(str.isalpha, col_reference.upper()))
        col_idx = column_index_from_string(col_letter)
        changes_made = 0
        files_processed = set()
        print("> Processing translations...")
        for row, data in all_sheets.items():
            for item in data:
                if item[col_idx - 1] is not None:
                    subdir = item[0] if item[0] else ""
                    filename = item[1] if item[1] else ""
                    msg_id = str(item[2]) if item[2] else ""
                    new_translation = item[col_idx - 1] if item[col_idx - 1] else ""
                    xml_filename = f"{filename}.pzd.xml"
                    xml_path = os.path.join(lang_root, subdir, xml_filename) if subdir else os.path.join(lang_root, xml_filename)
                    if not os.path.exists(xml_path):
                        print(f" \033[91m[ERROR]\033[00m File not found: {xml_path}")
                        continue
                    try:
                        with open(xml_path, "r", encoding="utf-8") as f:
                            content = f.read()
                        root = ET.fromstring(content)
                        if root.tag == "PzdFile":
                            root.set("xmlns:xsi", "http://www.w3.org/2001/XMLSchema-instance")
                            root.set("xmlns:xsd", "http://www.w3.org/2001/XMLSchema")
                        text_contents = root.find("TextContents")
                        if text_contents is not None:
                            for text_content in text_contents.findall("TextContent"):
                                if text_content.get("ID") == msg_id:
                                    message_elem = text_content.find("Message")
                                    if message_elem is not None:
                                        old_text = message_elem.text or ""
                                        message_elem.text = unescape(new_translation.strip())
                                        if old_text == new_translation:
                                            if verbose: print(f" \033[90m[SKIP] Message {msg_id} already translated, skipping.\033[00m")
                                            continue
                                        else:
                                            if verbose: print(f" \033[38;5;75m[INFO]\033[00m {filename} (ID: {msg_id}): \033[38;5;210m\"{old_text}\"\033[00m -> \033[38;5;81m\"{new_translation.strip()}\"\033[00m")
                                            changes_made += 1
                                            files_processed.add(xml_path)
                                        break
                        tree = ET.ElementTree(root)
                        write_xml(tree, xml_path)
                    except Exception as e:
                        print(f" \033[38;5;214m[WARNING]\033[00m Could not process {xml_path}: {e}")
                        continue
        print(f"\n \033[38;5;76m[DONE]\033[00m Summary:")
        print(f"   • {changes_made} translations applied.")
        print(f"   • {len(files_processed)} files modified.")
    except Exception as e:
        print(f" \033[91m[ERROR]\033[00m Error reading XLSX file: {e}")

# Command: convert-batch
def convert_batch(ff16converter, lang_path, valid_ext, verbose):
    from pathlib import Path
    from time import perf_counter
    from collections import defaultdict
    import subprocess
    lang_path = Path(lang_path)
    if not lang_path.exists():
        print(f" \033[91m[ERROR]\033[00m Folder {lang_path} does not exist")
        return
    if not Path(ff16converter).exists():
        print(f" \033[91m[ERROR]\033[00m Converter {ff16converter} does not exist")
        return
    if not valid_ext:
        print(f" \033[91m[ERROR]\033[00m Extension not set.")
        return
    valid_ext = [valid_ext]
    print(f"> Converting files in: \033[48;5;235m{lang_path}\033[00m\n> Processing. This may take a while...")
    files_to_convert = [
        file for file in lang_path.rglob("*")
        if file.suffix.lower() in valid_ext and file.is_file()
    ]
    if verbose: print(f" \033[38;5;75m[INFO]\033[00m {len(files_to_convert)} files to convert")
    folder_group = defaultdict(list); start_time = perf_counter()
    for file in files_to_convert:
        if valid_ext == [".pzd"]:
            has_xml = Path(str(file) + ".xml")
            if has_xml.exists():
                if verbose: print(f" \033[90m[SKIP] {has_xml.name} already exists, skipping.\033[00m")
                continue
            folder_group[str(file.parent.name)].append(file)
        else:
            has_pzd = Path(str(file) + "RB.pzd")
            if has_pzd.exists():
                if verbose: print(f" \033[90m[SKIP] {has_pzd.name} already exists, skipping.\033[00m")
                continue
            folder_group[str(file.parent.name)].append(file)
    for folder, files in folder_group.items():
        try:
            if verbose: print(f" \033[38;5;75m[INFO]\033[00m Converting files on: \033[38;5;81m{folder}\033[00m")
            if folder == "defaultq" or folder == "simpleq":
                helper = []
                for i in range(0, len(files), 400):
                    helper.append(files[i:i + 400])
                for chunk in helper:
                    subprocess.run([ff16converter] + [str(file) for file in chunk], capture_output=True, text=True)
            else:
                subprocess.run([ff16converter] + [str(file) for file in files], capture_output=True, text=True)
        except Exception as e:
            print(f" \033[91m[ERROR]\033[00m Error converting: {e}")
    time_lapsed = perf_counter() - start_time
    print(f" \033[38;5;76m[DONE]\033[00m Files converted in {int(time_lapsed // 3600):02d}:{int((time_lapsed % 3600) // 60):02d}:{int(time_lapsed % 60):02d}")

# Command: move-batch
def move_converted(this_directory, to_directory, extension, verbose):
    from pathlib import Path
    import shutil
    Path.mkdir(Path(to_directory), parents=True, exist_ok=True)
    print(f"> Moving files to: \033[48;5;235m{to_directory}\033[00m\n> Processing. This may take a while...")
    match extension:
        case ".xml":
            converted_files = list(Path(this_directory).rglob("*RB.pzd"))
            if not converted_files:
                return
            for file in converted_files:
                try:
                    relative_path = file.relative_to(this_directory)
                    original_name = file.name.replace(".pzd.xmlRB.pzd",".pzd")
                    destination_file = to_directory / relative_path.parent / original_name
                    Path.mkdir(destination_file.parent, parents=True, exist_ok=True)
                    if destination_file.exists():
                        if verbose: print(f" \033[90m[SKIP] {original_name} already exists, skipping.\033[00m")
                        continue
                    shutil.move(str(file), str(destination_file))
                    if verbose: print(f" \033[38;5;75m[INFO]\033[00m Moved: \033[38;5;81m{file.name}\033[00m to \033[48;5;235m{destination_file.parent}\033[00m")
                except Exception as e:
                    print(f" \033[91m[ERROR]\033[00m Error moving {file.name}: {e}")
        case ".pzd":
            converted_files = list(Path(this_directory).rglob("*.pzd.xml"))
            if not converted_files:
                return
            for file in converted_files:
                try:
                    relative_path = file.relative_to(this_directory)
                    destination_folder = to_directory / relative_path.parent
                    destination_file = destination_folder.parent / relative_path
                    Path.mkdir(destination_folder, parents=True, exist_ok=True)
                    if destination_file.exists():
                        if verbose: print(f" \033[90m[SKIP] {relative_path} already exists, skipping.\033[00m")
                        continue
                    shutil.move(str(file), str(destination_folder))
                    if verbose: print(f" \033[38;5;75m[INFO]\033[00m Moved: \033[38;5;81m{file.name}\033[00m to \033[48;5;235m{destination_folder}\033[00m")
                except Exception as e:
                    print(f" \033[91m[ERROR]\033[00m Error moving {file.name}: {e}")
    print(f" \033[38;5;76m[DONE]\033[00m Move operation completed.")

def main():
    os.system("color")
    parser = argparse.ArgumentParser(
        description="""\033[38;5;81m
 +----------------------------------------------+
 | FFXVI Subtitle Organizer v1.4                |
 | by Roysu                                     |
 +----------------------------------------------+
 | https://github.com/roymuke/FF16SubsOrganizer |
 +----------------------------------------------+\033[00m""",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
examples:
  \033[90m# Export subtitles to XLSX for editing\033[00m
  > \033[38;5;149mFF16SubsOrganizer.py\033[00m to-xlsx \033[38;5;149m-l\033[00m \033[38;5;222m"C:\path\\to\\folder\\0007.en.XML"\033[00m \033[38;5;149m-j\033[00m \033[38;5;222m"C:\path\\to\\folder\\0007.ja\\nxd\\txt"\033[00m \033[38;5;149m-o\033[00m \033[38;5;222m"C:\custom\path\\to\\file.xlsx"\033[00m

  \033[90m# Apply translations from XLSX back to XML\033[00m
  > \033[38;5;149mFF16SubsOrganizer.py\033[00m edit-xml \033[38;5;149m-f\033[00m \033[38;5;222m"C:\path\\to\\file.xlsx"\033[00m \033[38;5;149m-col\033[00m I2 \033[38;5;149m-l\033[00m \033[38;5;222m"C:\path\\to\\folder\\0007.en.XML"\033[00m

  \033[90m# Convert in batch PZD->XML or XML->PZD\033[00m
  > \033[38;5;149mFF16SubsOrganizer.py\033[00m convert-batch \033[38;5;149m-c\033[00m \033[38;5;222m"C:\path\\to\\FF16Converter.exe"\033[00m \033[38;5;149m-f\033[00m \033[38;5;222m"C:\path\\to\\folder\\0007.en\\nxd\\text"\033[00m \033[38;5;149m--pzd\033[00m \033[38;5;149m-m\033[00m \033[38;5;222m"C:\path\\to\moving\\folder"\033[00m

  \033[90m# Move files to another destination by extension\033[00m
  > \033[38;5;149mFF16SubsOrganizer.py\033[00m move-batch \033[38;5;149m-f\033[00m \033[38;5;222m"C:\path\\to\\folder\\0007.en.XML"\033[00m \033[38;5;149m--pzd\033[00m \033[38;5;149m-m\033[00m \033[38;5;222m"C:\path\\to\\folder\\0007.en.PZD"\033[00m""")
    subparsers = parser.add_subparsers(dest="command", required=True, help="Available commands")
    # to-xlsx command
    xlsx_parser = subparsers.add_parser("to-xlsx", help="Export subtitles to XLSX file.")
    xlsx_parser.add_argument("-l", "--language", required=True, help="Path to language subs folder to translate")
    xlsx_parser.add_argument("-j", "--japanese", required=True, help="Path to Japanese  subsfolder")
    xlsx_parser.add_argument("-o", "--output", default="ff16_subtitles.xlsx", help="Output xlsx file")
    xlsx_parser.add_argument("-v", "--verbose", action="store_true", help="Show detailed output messages")
    # edit-xml command
    edit_parser = subparsers.add_parser("edit-xml", help="Gets translations from XLSX back to XML files.")
    edit_parser.add_argument("-f", "--file", required=True, help="XLSX file path")
    edit_parser.add_argument("-col", required=True, help="Column with new translations (e.g. I2)")
    edit_parser.add_argument("-l", "--language", required=True, help="Path to language to translate folder (e.g. C:\...\0007.en\nxd\text)")
    edit_parser.add_argument("-v", "--verbose", action="store_true", help="Show detailed output messages")
    # convert-batch command
    batch_parser = subparsers.add_parser("convert-batch", help="Convert files to another format, pzd->xml OR xml->pzd.")
    batch_parser.add_argument("-c", "--converter", required=True, help="Path to FF16Converter.exe")
    batch_parser.add_argument("-f", "--folder", required=True, help="Path to language folder")
    batch_parser.add_argument("--pzd", action="store_const", const=".pzd", dest="extension", help="Extension to convert (pzd -> xml).")
    batch_parser.add_argument("--xml", action="store_const", const=".xml", dest="extension", help="Extension to convert (xml -> pzd).")
    batch_parser.add_argument("-m", "--moveto", help="Path to converted files folder destination.")
    batch_parser.add_argument("-v", "--verbose", action="store_true", help="Show detailed output messages")
    # move-batch command
    move_parser = subparsers.add_parser("move-batch", help="Move files to another destination.")
    move_parser.add_argument("-f", "--folder", required=True, help="Path to parent folder.")
    move_parser.add_argument("--pzd", action="store_const", const=".pzd", dest="extension", help="Move PZD files.")
    move_parser.add_argument("--xml", action="store_const", const=".xml", dest="extension", help="Move XML files.")
    move_parser.add_argument("-m", "--moveto", required=True, help="Path to folder destination.")
    move_parser.add_argument("-v", "--verbose", action="store_true", help="Show detailed output messages")

    args = parser.parse_args()

    if args.command == "to-xlsx":
        print(f"> Exporting to XLSX: \033[48;5;235m{args.output}\033[00m")
        table_rows = collect_table(args.language, args.japanese)
        export_xlsx(table_rows, args.output, args.verbose)
    elif args.command == "edit-xml":
        print(f"> Applying translations from: \033[48;5;235m{args.file}\033[00m")
        edit_xml(args.file, args.col, args.language, args.verbose)
    elif args.command == "convert-batch":
        convert_batch(args.converter,args.folder,args.extension, args.verbose)
        if args.moveto and args.extension:
            move_converted(args.folder, args.moveto, args.extension, args.verbose)
    elif args.command == "move-batch":
        move_converted(args.folder, args.moveto, args.extension, args.verbose)

if __name__ == "__main__":
    main()